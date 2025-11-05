from odoo import models, fields, api
from datetime import date, timedelta
import io, base64
from openpyxl import Workbook
from openpyxl.styles import Font
import logging

_logger = logging.getLogger(__name__)

class AccountMove(models.Model):
    _inherit = 'account.move'

    lead_id = fields.Many2one(
        'crm.lead',
        string='Lead',
        help="The lead associated with this invoice.",
    )

    @api.model
    def _generate_and_send_account_report(self, frequency):
        """Generate daily or weekly accounting summary report and send to configured emails."""
        today = date.today()
        start_date = today - timedelta(days=7) if frequency == 'Weekly' else today

        # Common domain
        domain_period = [
            ('invoice_date', '>=', start_date),
            ('invoice_date', '<=', today),
            ('state', '=', 'posted')
        ]

        # Credit Invoices
        credit_invoices = self.search([('move_type', '=', 'out_refund'), *domain_period])
        credit_count = len(credit_invoices)
        credit_sum = sum(credit_invoices.mapped('amount_total'))

        # Payments Collected
        payments = self.env['account.payment'].search([
            ('ate', '>=', start_date),
            ('date', '<=', today),
            ('state', 'in', ['posted', 'in_process'])
        ])
        payment_count = len(payments)
        payment_sum = sum(payments.mapped('amount'))

        # Payments Due
        due_invoices = self.search([
            ('move_type', '=', 'out_invoice'),
            ('payment_state', 'in', ['not_paid', 'partial']),
            ('state', '=', 'posted')
        ])
        due_count = len(due_invoices)
        due_sum = sum(due_invoices.mapped('amount_residual'))

        # Return Payments (Refunds marked as Return)
        return_payments = self.search([
            ('move_type', '=', 'out_refund'),
            ('invoice_origin', 'ilike', 'Return'),
            *domain_period
        ])
        return_count = len(return_payments)
        return_sum = sum(return_payments.mapped('amount_total'))

        # Create Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = f"{frequency} Report {today}"

        headers = ["Section", "Count", "Total Amount"]
        ws.append(headers)
        ws.append(["Credit Invoices", credit_count, credit_sum])
        ws.append(["Payments Collected", payment_count, payment_sum])
        ws.append(["Payments Due", due_count, due_sum])
        ws.append(["Return Payments", return_count, return_sum])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        fp = io.BytesIO()
        if ws.max_row > 1:
            wb.save(fp)
        else:
            fp.close()
            _logger.info('No data rows after headers; Excel report not generated.')
            return
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()

        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': f"{frequency}_Account_Report_{today}.xlsx",
            'type': 'binary',
            'datas': file_data,
            'res_model': 'account.move',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Get recipient emails (comma-separated)
        recipient_param = self.env['ir.config_parameter'].sudo().get_param('account.report_email')
        if not recipient_param:
            _logger.warning("⚠️ No email configured in settings (account.report_email).")
            return

        recipients = [email.strip() for email in recipient_param.split(',') if email.strip()]
        email_to = ','.join(recipients)

        if recipients:
            mail_values = {
                'email_from': self.env.user.email_formatted or 'no-reply@cultstay.com',
                'subject': f'{frequency} Accounting Report - {today}',
                'body_html': f'<p>Hello,</p><p>Attached is the {frequency.lower()} accounting summary report.</p>',
                'email_to': email_to,
                'attachment_ids': [(6, 0, [attachment.id])],
            }
            mail = self.env['mail.mail'].create(mail_values)
            mail.send()
            _logger.info("✅ %s report sent to %s", frequency, email_to)
        else:
            _logger.warning("⚠️ No valid email addresses found in settings.")

    def _send_daily_credit_report(self):
        today = date.today()

        # Common domain
        domain_period = [
            ('invoice_date', '=', today),
            ('state', '=', 'posted')
        ]

        # Credit Invoices
        wb = Workbook()
        ws = wb.active
        ws.title = f"Daily Credit Report {today}"
        headers = ["Invoice Number", "Customer", "Invoice Date", "Total Amount"]
        ws.append(headers)
        credit_invoices = self.search([('move_type', '=', 'out_refund'), *domain_period])
        for invoice in credit_invoices:
            ws.append([
                invoice.name,
                invoice.partner_id.name,
                invoice.invoice_date,
                invoice.amount_total
            ])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        fp = io.BytesIO()
        if ws.max_row > 1:
            wb.save(fp)
        else:
            fp.close()
            _logger.info('No data rows after headers; Excel report not generated.')
            return
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()
        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': f"Daily_Credit_Report_{today}.xlsx",
            'type': 'binary',
            'datas': file_data,
            'res_model': 'account.move',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })

        # Get recipient emails (comma-separated)
        recipient_param = self.env['ir.config_parameter'].sudo().get_param('account.report_email')
        if not recipient_param:
            _logger.warning("⚠️ No email configured in settings (account.report_email).")
            return
        recipients = [email.strip() for email in recipient_param.split(',') if email.strip()]
        email_to = ','.join(recipients)
        if recipients:
            mail_values = {
                'email_from': self.env.user.email_formatted,
                'subject': f'Daily Credit Report - {today}',
                'body_html': f'<p>Hello,</p><p>Attached is the daily credit invoice report.</p>',
                'email_to': email_to,
                'attachment_ids': [(6, 0, [attachment.id])],
            }
            mail = self.env['mail.mail'].create(mail_values)
            mail.send()
            _logger.info("✅ Daily credit report sent to %s", email_to)
        else:
            _logger.warning("⚠️ No valid email addresses found in settings.")

    def _send_daily_payment_report(self):
        today = date.today()
        domain_period = [
            ('date', '=', today),
            ('state', '=', 'paid')
        ]
        payments = self.env['account.payment'].search(domain_period)
        if not payments:
            _logger.info("No payments found for today; report not generated.")
            return
        wb = Workbook()
        ws = wb.active
        ws.title = f"Daily Payment Report {today}"
        headers = ["Payment Reference", "Customer", "Phone", "Payment Date", "Amount"]
        ws.append(headers)
        for payment in payments:
            ws.append([
                payment.name,
                payment.partner_id.name,
                payment.partner_id.phone,
                payment.date,
                payment.amount
            ])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        fp = io.BytesIO()
        if ws.max_row > 1:
            wb.save(fp)
        else:
            fp.close()
            _logger.info('No data rows after headers; Excel report not generated.')
            return
        fp.seek(0)
        file_data = base64.b64encode(fp.read())
        fp.close()
        # Create attachment
        attachment = self.env['ir.attachment'].create({
            'name': f"Daily_Payment_Report_{today}.xlsx",
            'type': 'binary',
            'datas': file_data,
            'res_model': 'account.move',
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
    
            
