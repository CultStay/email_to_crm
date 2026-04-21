# -*- coding: utf-8 -*-
import io
import base64
from datetime import date, timedelta
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
except ImportError:
    raise UserError("openpyxl library is required. Install via: pip install openpyxl")


class CollectionReportWizard(models.TransientModel):
    _name = 'collection.report.wizard'
    _description = 'Collection Report Wizard'

    # ── Date Options ────────────────────────────────────────────────────────
    date_filter = fields.Selection([
        ('14', '14 Days (2 Weeks)'),
        ('21', '21 Days (3 Weeks)'),
        ('28', '28 Days (4 Weeks)'),
        ('custom', 'Custom Date Range'),
    ], string='Report Period', required=True, default='14')

    date_from = fields.Date(
        string='Date From',
        default=lambda self: date.today() - timedelta(days=14)
    )
    date_to = fields.Date(
        string='Date To',
        default=fields.Date.today
    )

    city = fields.Char('City')

    # ── Filters ──────────────────────────────────────────────────────────────
    # city_ids = fields.Many2many(
    #     'res.city', string='Cities',
    #     help="Leave empty to include all cities"
    # )
    company_ids = fields.Many2many(
        'res.company', string='Companies',
        default=lambda self: self.env.companies
    )
    include_draft = fields.Boolean(string='Include Draft Invoices', default=False)

    # ── Computed label ───────────────────────────────────────────────────────
    period_label = fields.Char(compute='_compute_period_label')

    @api.depends('date_filter', 'date_from', 'date_to')
    def _compute_period_label(self):
        for rec in self:
            if rec.date_filter == 'custom':
                rec.period_label = f"{rec.date_from} → {rec.date_to}"
            else:
                rec.period_label = f"Last {rec.date_filter} Days"

    @api.onchange('date_filter')
    def _onchange_date_filter(self):
        today = date.today()
        if self.date_filter != 'custom':
            days = int(self.date_filter)
            self.date_from = today - timedelta(days=days)
            self.date_to = today

    # ── Main Action ──────────────────────────────────────────────────────────
    def action_generate_report(self):
        self.ensure_one()

        # Validate dates
        if self.date_from > self.date_to:
            raise UserError(_("Start date cannot be after end date!"))

        invoices = self._fetch_invoices()
        if not invoices:
            raise UserError(
                _("No invoices found for the selected criteria.")
            )
        xlsx_data = self._build_excel(invoices)
        filename = f"Collection_Report_{self.date_from}_{self.date_to}.xlsx"

        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ── Data Fetch ────────────────────────────────────────────────────────────
    def _fetch_invoices(self):
        domain = [
            ('move_type', '=', 'out_invoice'),
            ('create_date', '>=', self.date_from),
            ('create_date', '<=', self.date_to),
        ]
        # if not self.include_draft:
        #     domain.append(('state', '=', 'posted'))
        if self.company_ids:
            domain.append(('company_id', 'in', self.company_ids.ids))

        # Apply city filter
        if self.city:
            domain.append(('partner_id.city', 'ilike', self.city.strip()))

        invoices = self.env['account.move'].search(domain, order='invoice_date asc')

        # if self.city_ids:
        #     invoices = invoices.filtered(
        #         lambda inv: inv.partner_id.city_id.id in self.city_ids.ids
        #         if hasattr(inv.partner_id, 'city_id') else
        #         inv.partner_id.city in [c.name for c in self.city_ids]
        #     )
        return invoices

    def _get_sale_orders_from_invoice(self, invoice):
        """Get sale order references from invoice lines"""
        sale_orders = invoice.invoice_line_ids.mapped('sale_line_ids.order_id')
        return sale_orders.filtered(lambda so: so.state not in ('cancel', 'draft'))

    # ── Excel Builder ─────────────────────────────────────────────────────────
    def _build_excel(self, invoices):
        wb = openpyxl.Workbook()
        today = date.today()

        # ── Sheet 1: Summary by City ────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary by City"
        self._write_summary_sheet(ws_summary, invoices, today)

        # ── Sheet 2: Full Detail ────────────────────────────────────────────
        ws_detail = wb.create_sheet("Invoice Details")
        self._write_detail_sheet(ws_detail, invoices, today)

        # ── Sheet 3: Aging Analysis ─────────────────────────────────────────
        ws_aging = wb.create_sheet("Aging Analysis")
        self._write_aging_sheet(ws_aging, invoices, today)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Styles Helper ─────────────────────────────────────────────────────────
    @staticmethod
    def _styles():
        thin = Side(style='thin', color='BDBDBD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        return {
            'title': Font(name='Arial', bold=True, size=14, color='FFFFFF'),
            'header': Font(name='Arial', bold=True, size=10, color='FFFFFF'),
            'body': Font(name='Arial', size=9),
            'body_bold': Font(name='Arial', bold=True, size=9),
            'title_fill': PatternFill('solid', start_color='1F3864'),
            'header_fill': PatternFill('solid', start_color='2E75B6'),
            'alt_fill': PatternFill('solid', start_color='EBF3FB'),
            'bucket_14': PatternFill('solid', start_color='C6EFCE'),   # green
            'bucket_21': PatternFill('solid', start_color='FFEB9C'),   # yellow
            'bucket_28': PatternFill('solid', start_color='FFC7CE'),   # red
            'total_fill': PatternFill('solid', start_color='D6DCE4'),
            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left': Alignment(horizontal='left', vertical='center'),
            'right': Alignment(horizontal='right', vertical='center'),
            'border': border,
        }

    def _write_summary_sheet(self, ws, invoices, today):
        s = self._styles()
        # Title Row
        ws.merge_cells('A1:G1')
        ws['A1'] = f"COLLECTION REPORT — {self.period_label}"
        ws['A1'].font = s['title']
        ws['A1'].fill = s['title_fill']
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 30

        ws.merge_cells('A2:G2')
        ws['A2'] = f"Generated: {today}  |  Period: {self.date_from} to {self.date_to}"
        ws['A2'].font = Font(name='Arial', italic=True, size=9, color='555555')
        ws['A2'].alignment = s['center']

        # Headers
        headers = ['City', 'Total Invoices', 'Total Amount', '14-Day Amount', '21-Day Amount', '28-Day Amount', 'Overdue %']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = s['header']
            cell.fill = s['header_fill']
            cell.alignment = s['center']
            cell.border = s['border']
        ws.row_dimensions[3].height = 20

        # Group by city
        city_data = {}
        for inv in invoices:
            city = inv.partner_id.city or ''
            age = (today - inv.invoice_date).days if inv.invoice_date else 0
            amt = inv.amount_total_signed
            if city not in city_data:
                city_data[city] = {'count': 0, 'total': 0, 'b14': 0, 'b21': 0, 'b28': 0}
            city_data[city]['count'] += 1
            city_data[city]['total'] += amt
            if age >= 28:
                city_data[city]['b28'] += amt
            elif age >= 21:
                city_data[city]['b21'] += amt
            elif age >= 14:
                city_data[city]['b14'] += amt

        row = 4
        for i, (city, d) in enumerate(sorted(city_data.items())):
            fill = s['alt_fill'] if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
            overdue_pct = ((d['b14'] + d['b21'] + d['b28']) / d['total'] * 100) if d['total'] else 0
            vals = [city, d['count'], d['total'], d['b14'], d['b21'], d['b28'], overdue_pct / 100]
            aligns = [s['left'], s['center'], s['right'], s['right'], s['right'], s['right'], s['center']]
            for col, (v, al) in enumerate(zip(vals, aligns), 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.font = s['body']
                cell.fill = fill
                cell.alignment = al
                cell.border = s['border']
                if col in (3, 4, 5, 6):
                    cell.number_format = '#,##0.00'
                if col == 7:
                    cell.number_format = '0.0%'
            row += 1

        # Totals
        total_row = row
        ws.cell(total_row, 1, 'TOTAL').font = s['body_bold']
        ws.cell(total_row, 1).fill = s['total_fill']
        ws.cell(total_row, 1).border = s['border']
        ws.cell(total_row, 1).alignment = s['left']
        for col in range(2, 8):
            cell = ws.cell(total_row, col)
            cell.font = s['body_bold']
            cell.fill = s['total_fill']
            cell.border = s['border']
            if col == 2:
                cell.value = f'=SUM(B4:B{total_row-1})'
                cell.alignment = s['center']
            elif col in (3, 4, 5, 6):
                col_letter = get_column_letter(col)
                cell.value = f'=SUM({col_letter}4:{col_letter}{total_row-1})'
                cell.number_format = '#,##0.00'
                cell.alignment = s['right']
            elif col == 7:
                cell.value = f'=IF(C{total_row}=0,0,(D{total_row}+E{total_row}+F{total_row})/C{total_row})'
                cell.number_format = '0.0%'
                cell.alignment = s['center']

        # Column widths
        widths = [22, 15, 18, 18, 18, 18, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Legend
        ws.cell(row + 2, 1, 'Legend:').font = s['body_bold']
        for col, (label, fill_key) in enumerate([
            ('14-Day (≥14 days old)', 'bucket_14'),
            ('21-Day (≥21 days old)', 'bucket_21'),
            ('28-Day (≥28 days old)', 'bucket_28'),
        ], 2):
            c = ws.cell(row + 2, col, label)
            c.fill = s[fill_key]
            c.font = s['body']
            c.alignment = s['center']
            c.border = s['border']

    def _write_detail_sheet(self, ws, invoices, today):
        s = self._styles()
        ws.merge_cells('A1:K1')
        ws['A1'] = "INVOICE DETAILS — COLLECTION REPORT"
        ws['A1'].font = s['title']
        ws['A1'].fill = s['title_fill']
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 28

        headers = [
            'Invoice No.', 'Invoice Date', 'Due Date', 'Customer',
            'City', 'Salesperson', 'Sale Order Ref',
            'Invoice Amount', 'Amount Due', 'Days Old', 'Aging Bucket'
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(2, col, h)
            cell.font = s['header']
            cell.fill = s['header_fill']
            cell.alignment = s['center']
            cell.border = s['border']
        ws.row_dimensions[2].height = 18

        for i, inv in enumerate(invoices):
            row = i + 3
            age = (today - inv.invoice_date).days if inv.invoice_date else 0
            if age >= 28:
                bucket = '28+ Days'
                bfill = s['bucket_28']
            elif age >= 21:
                bucket = '21 Days'
                bfill = s['bucket_21']
            elif age >= 14:
                bucket = '14 Days'
                bfill = s['bucket_14']
            else:
                bucket = '< 14 Days'
                bfill = PatternFill('solid', start_color='FFFFFF')

            base_fill = s['alt_fill'] if i % 2 == 0 else PatternFill('solid', start_color='FFFFFF')
            city = inv.partner_id.city or ''
            sale_refs = ', '.join(inv.invoice_line_ids.mapped('sale_line_ids.order_id.name')) or ''
            salesperson = inv.invoice_user_id.name or inv.user_id.name or ''

            row_data = [
                (inv.name, s['left']),
                (inv.invoice_date, s['center']),
                (inv.invoice_date_due, s['center']),
                (inv.partner_id.name, s['left']),
                (city, s['left']),
                (salesperson, s['left']),
                (sale_refs, s['left']),
                (inv.amount_total, s['right']),
                (inv.amount_residual, s['right']),
                (age, s['center']),
                (bucket, s['center']),
            ]
            for col, (val, align) in enumerate(row_data, 1):
                cell = ws.cell(row, col, val)
                cell.font = s['body']
                cell.fill = bfill if col == 11 else base_fill
                cell.alignment = align
                cell.border = s['border']
                if col in (2, 3) and val:
                    cell.number_format = 'DD/MM/YYYY'
                if col in (8, 9):
                    cell.number_format = '#,##0.00'

        # Autofit column widths
        col_widths = [18, 13, 13, 28, 16, 18, 18, 16, 14, 10, 13]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:K{len(invoices)+2}'

    def _write_aging_sheet(self, ws, invoices, today):
        s = self._styles()
        ws.merge_cells('A1:E1')
        ws['A1'] = "AGING ANALYSIS — 14 / 21 / 28 DAYS"
        ws['A1'].font = s['title']
        ws['A1'].fill = s['title_fill']
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 28

        # Bucket breakdown
        buckets = {'< 14 Days': [], '14 Days': [], '21 Days': [], '28+ Days': []}
        for inv in invoices:
            age = (today - inv.invoice_date).days if inv.invoice_date else 0
            if age >= 28:
                buckets['28+ Days'].append(inv)
            elif age >= 21:
                buckets['21 Days'].append(inv)
            elif age >= 14:
                buckets['14 Days'].append(inv)
            else:
                buckets['< 14 Days'].append(inv)

        bucket_fills = {
            '< 14 Days': PatternFill('solid', start_color='FFFFFF'),
            '14 Days': s['bucket_14'],
            '21 Days': s['bucket_21'],
            '28+ Days': s['bucket_28'],
        }

        for col, h in enumerate(['Aging Bucket', 'No. of Invoices', 'Total Amount', 'Amount Due', '% of Total'], 1):
            cell = ws.cell(2, col, h)
            cell.font = s['header']
            cell.fill = s['header_fill']
            cell.alignment = s['center']
            cell.border = s['border']

        grand_total = sum(inv.amount_total for inv in invoices)
        row = 3
        for bucket, invs in buckets.items():
            total_amt = sum(i.amount_total for i in invs)
            due_amt = sum(i.amount_residual for i in invs)
            pct = (total_amt / grand_total) if grand_total else 0
            vals = [bucket, len(invs), total_amt, due_amt, pct]
            aligns = [s['left'], s['center'], s['right'], s['right'], s['center']]
            for col, (v, al) in enumerate(zip(vals, aligns), 1):  # ← zip here
                cell = ws.cell(row, col, v)
                cell.font = s['body_bold'] if col == 1 else s['body']
                cell.fill = bucket_fills[bucket]
                cell.alignment = al
                cell.border = s['border']
                if col in (3, 4):
                    cell.number_format = '#,##0.00'
                if col == 5:
                    cell.number_format = '0.0%'
            row += 1

        # Grand Total row
        for col in range(1, 6):
            cell = ws.cell(row, col)
            cell.font = s['body_bold']
            cell.fill = s['total_fill']
            cell.border = s['border']
        ws.cell(row, 1, 'GRAND TOTAL').alignment = s['left']
        ws.cell(row, 2, f'=SUM(B3:B{row-1})').alignment = s['center']
        ws.cell(row, 3, f'=SUM(C3:C{row-1})').number_format = '#,##0.00'
        ws.cell(row, 3).alignment = s['right']
        ws.cell(row, 4, f'=SUM(D3:D{row-1})').number_format = '#,##0.00'
        ws.cell(row, 4).alignment = s['right']
        ws.cell(row, 5, '100.0%').number_format = '0.0%'
        ws.cell(row, 5).alignment = s['center']

        for i, w in enumerate([18, 16, 18, 16, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
