# -*- coding: utf-8 -*-
import io
import base64
from datetime import date, timedelta
from collections import defaultdict
from odoo import models, fields, api, _
from odoo.exceptions import UserError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise UserError("openpyxl library is required. Install via: pip install openpyxl")


class ReturnReportWizard(models.TransientModel):
    _name = 'return.report.wizard'
    _description = 'Return Report Wizard'

    # ── Date Options ────────────────────────────────────────────────────────
    date_from = fields.Date(
        string='Date From',
        required=True,
        default=lambda self: date.today().replace(day=1),
    )
    date_to = fields.Date(
        string='Date To',
        required=True,
        default=fields.Date.today,
    )

    # ── Filters ──────────────────────────────────────────────────────────────
    company_ids = fields.Many2many(
        'res.company',
        string='Companies',
        default=lambda self: self.env.companies,
    )
    partner_ids = fields.Many2many(
        'res.partner',
        string='Shops / Customers',
        help='Leave empty to include all shops/customers.',
    )
    product_ids = fields.Many2many(
        'product.product',
        string='Products',
        help='Leave empty to include all products.',
    )

    # ── Computed label ───────────────────────────────────────────────────────
    period_label = fields.Char(compute='_compute_period_label')

    @api.depends('date_from', 'date_to')
    def _compute_period_label(self):
        for rec in self:
            rec.period_label = f"{rec.date_from} → {rec.date_to}"

    # ── Main Action ──────────────────────────────────────────────────────────
    def action_generate_report(self):
        self.ensure_one()
        data = self._fetch_data()
        if not data['invoices'] and not data['returns']:
            raise UserError(_("No invoices or return records found for the selected criteria."))
        xlsx_data = self._build_excel(data)
        filename = f"Return_Report_{self.date_from}_{self.date_to}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(xlsx_data),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    # ── Data Fetch ────────────────────────────────────────────────────────────
    def _fetch_data(self):
        """
        Returns a dict with:
          invoices – account.move (out_invoice) in date range
          returns  – account.move (out_refund / credit note) in date range
                     PLUS stock return qty via sale order line quantities
        """
        base_domain = [
            ('invoice_date', '>=', self.date_from),
            ('invoice_date', '<=', self.date_to),
            ('state', '=', 'posted'),
        ]
        if self.company_ids:
            base_domain.append(('company_id', 'in', self.company_ids.ids))

        inv_domain = base_domain + [('move_type', '=', 'out_invoice')]
        ret_domain = base_domain + [('move_type', '=', 'out_refund')]

        if self.partner_ids:
            inv_domain.append(('partner_id', 'in', self.partner_ids.ids))
            ret_domain.append(('partner_id', 'in', self.partner_ids.ids))

        invoices = self.env['account.move'].search(inv_domain, order='invoice_date asc, partner_id asc')
        returns  = self.env['account.move'].search(ret_domain, order='invoice_date asc, partner_id asc')
        return {'invoices': invoices, 'returns': returns}

    # ── Return Qty Detection ─────────────────────────────────────────────────
    # def _build_return_map(self, invoices, credit_notes, wanted_products):
    #     """
    #     Return qty has TWO sources in Odoo — we combine both:
    #
    #     Source 1 — Credit notes (out_refund):
    #         Customer explicitly credited/refunded via Accounting.
    #
    #     Source 2 — Stock delivery returns (stock.picking type=incoming
    #         originated from a customer delivery):
    #         When the user does Delivery → Return, Odoo creates a reverse
    #         picking. The returned qty on the sale order line =
    #             product_uom_qty (original SO line)
    #             MINUS qty_delivered (net delivered after stock return)
    #         This difference = qty returned via warehouse, never invoiced.
    #
    #     We key by (invoice_id, product_id) so each invoice row gets its
    #     own accurate return qty rather than mixing across dates.
    #     """
    #     return_map = defaultdict(lambda: {'qty': 0.0, 'amt': 0.0})
    #
    #     # ── Source 1: Credit notes ───────────────────────────────────────────
    #     for ret in credit_notes:
    #         # Try to link back to the original invoice
    #         origin_inv_ids = ret.reversed_entry_id.ids if ret.reversed_entry_id else []
    #         for line in ret.invoice_line_ids.filtered(
    #             lambda l: l.product_id and l.display_type not in ('line_section', 'line_note')
    #         ):
    #             if wanted_products and line.product_id.id not in wanted_products:
    #                 continue
    #             # Key on originating invoice if known, else partner+date
    #             if origin_inv_ids:
    #                 for inv_id in origin_inv_ids:
    #                     key = (inv_id, line.product_id.id)
    #                     return_map[key]['qty'] += line.quantity
    #                     return_map[key]['amt'] += abs(line.price_subtotal)
    #             else:
    #                 # Fallback: key on (partner_id, invoice_date, product_id)
    #                 key = ('partner', ret.partner_id.id, ret.invoice_date, line.product_id.id)
    #                 return_map[key]['qty'] += line.quantity
    #                 return_map[key]['amt'] += abs(line.price_subtotal)
    #
    #     # ── Source 2: Stock delivery returns via sale order lines ────────────
    #     # For every invoice line, trace back to the sale order line and compare
    #     # product_uom_qty (qty ordered on SO) vs qty_delivered (net after returns)
    #     for inv in invoices:
    #         for inv_line in inv.invoice_line_ids.filtered(
    #             lambda l: l.product_id and l.display_type not in ('line_section', 'line_note')
    #         ):
    #             if wanted_products and inv_line.product_id.id not in wanted_products:
    #                 continue
    #             prod_id = inv_line.product_id.id
    #
    #             # Walk through linked sale order lines
    #             for sol in inv_line.sale_line_ids:
    #                 # qty_ordered = what was originally ordered on the SO line
    #                 # qty_delivered = what was actually net-delivered (after stock returns)
    #                 # stock_return_qty = difference
    #                 stock_return_qty = max(sol.product_uom_qty - sol.qty_delivered, 0)
    #                 if stock_return_qty > 0:
    #                     key = (inv.id, prod_id)
    #                     # Use unit price from invoice line for amount
    #                     unit_price = inv_line.price_unit * (1 - inv_line.discount / 100.0)
    #                     return_map[key]['qty'] = max(
    #                         return_map[key]['qty'], stock_return_qty
    #                     )
    #                     return_map[key]['amt'] = max(
    #                         return_map[key]['amt'], stock_return_qty * unit_price
    #                     )
    #
    #     return return_map

    def _build_return_map(self, invoices, credit_notes, wanted_products):
        """
        Calculate return quantities from STOCK RETURN PICKINGS (stock.move)

        Logic:
        1. Find all sale orders linked to invoices
        2. Find all return pickings (stock.picking with type='incoming' and
           origin that references a delivery)
        3. Match return moves to original delivery moves
        4. Calculate returned quantity per (invoice_id, product_id)

        Returns:
            return_map: dict with key (invoice_id, product_id) containing
                       {'qty': float, 'amt': float}
        """
        return_map = defaultdict(lambda: {'qty': 0.0, 'amt': 0.0})

        # Get all sale orders from invoices
        sale_orders = self.env['sale.order']
        invoice_sale_map = {}  # invoice_id -> sale_order

        for inv in invoices:
            # Get sale orders from invoice lines
            for inv_line in inv.invoice_line_ids:
                if inv_line.sale_line_ids:
                    so = inv_line.sale_line_ids[0].order_id
                    if so not in sale_orders:
                        sale_orders |= so
                    invoice_sale_map[inv.id] = so

        if not sale_orders:
            return return_map

        # Find all return pickings linked to these sale orders
        # Return pickings have:
        # - picking_type_id.code = 'incoming' (return to warehouse)
        # - origin contains original delivery name or sale order name
        # - backorder_id points to original delivery (or we can trace through moves)

        return_pickings = self.env['stock.picking'].search([
            ('origin', 'in', sale_orders.mapped('name')),
            ('picking_type_id.code', '=', 'incoming'),  # Return type
            ('state', '=', 'done'),
            ('date_done', '>=', self.date_from),
            ('date_done', '<=', self.date_to),
        ])

        # Also search by sale_order_id directly if available
        if hasattr(self.env['stock.picking'], 'sale_id'):
            return_pickings |= self.env['stock.picking'].search([
                ('sale_id', 'in', sale_orders.ids),
                ('picking_type_id.code', '=', 'incoming'),
                ('state', '=', 'done'),
                ('date_done', '>=', self.date_from),
                ('date_done', '<=', self.date_to),
            ])

        # Map return moves to original delivery moves
        for return_picking in return_pickings:
            # Get all return moves (positive quantity = returning stock)
            return_moves = return_picking.move_ids.filtered(
                lambda m: m.product_id and m.state == 'done' and m.quantity > 0
            )

            for return_move in return_moves:
                if wanted_products and return_move.product_id.id not in wanted_products:
                    continue

                # Find the original delivery move
                # Method 1: Check move_orig_ids (original move that created this return)
                original_move = False
                if return_move.move_orig_ids:
                    original_move = return_move.move_orig_ids[0]
                # Method 2: Search by product and sale line
                else:
                    # Find the original delivery move from the same sale order
                    sale_order = return_picking.sale_id or return_move.sale_line_id.order_id
                    if sale_order:
                        # Find delivery pickings for this sale order
                        delivery_pickings = sale_order.picking_ids.filtered(
                            lambda p: p.picking_type_id.code == 'outgoing' and p.state == 'done'
                        )
                        for delivery in delivery_pickings:
                            delivery_move = delivery.move_ids.filtered(
                                lambda m: m.product_id == return_move.product_id and
                                          m.state == 'done' and
                                          m.quantity > 0
                            )
                            if delivery_move:
                                original_move = delivery_move[0]
                                break

                if original_move:
                    # Find which invoice this original move was invoiced on
                    invoice_line = self.env['account.move.line'].search([
                        ('sale_line_ids', 'in', original_move.sale_line_id.ids),
                        ('move_id.move_type', '=', 'out_invoice'),
                        ('move_id.state', '=', 'posted'),
                    ], limit=1)

                    if invoice_line and invoice_line.move_id:
                        invoice = invoice_line.move_id
                        prod_id = return_move.product_id.id
                        key = (invoice.id, prod_id)

                        return_qty = return_move.quantity
                        # Use the original invoice line price for amount
                        inv_line = invoice.invoice_line_ids.filtered(
                            lambda l: l.product_id.id == prod_id
                        )
                        if inv_line:
                            unit_price = inv_line[0].price_unit * (1 - inv_line[0].discount / 100.0)
                            return_map[key]['qty'] += return_qty
                            return_map[key]['amt'] += return_qty * unit_price
                        else:
                            return_map[key]['qty'] += return_qty
                            return_map[key]['amt'] += return_qty * return_move.price_unit

        # Also process credit notes (if needed for financial returns)
        for ret in credit_notes:
            origin_inv_ids = ret.reversed_entry_id.ids if ret.reversed_entry_id else []
            for line in ret.invoice_line_ids.filtered(
                    lambda l: l.product_id and l.display_type not in ('line_section', 'line_note')
            ):
                if wanted_products and line.product_id.id not in wanted_products:
                    continue
                if origin_inv_ids:
                    for inv_id in origin_inv_ids:
                        key = (inv_id, line.product_id.id)
                        # Only add credit note return if it's not already counted in stock returns
                        # Credit notes might be for non-stock items or service returns
                        if return_map[key]['qty'] < abs(line.quantity):
                            additional_qty = abs(line.quantity) - return_map[key]['qty']
                            return_map[key]['qty'] += additional_qty
                            return_map[key]['amt'] += abs(line.price_subtotal)
                else:
                    # For credit notes without original invoice reference
                    key = ('partner', ret.partner_id.id, ret.invoice_date, line.product_id.id)
                    return_map[key]['qty'] += abs(line.quantity)
                    return_map[key]['amt'] += abs(line.price_subtotal)

        return return_map

    # ── Data Aggregation ──────────────────────────────────────────────────────
    def _aggregate(self, data):
        """
        Build a nested structure:
          result[invoice_date][partner_id] = {
              'partner_name': ...,
              'invoice_name': ...,
              'invoice_id': ...,
              'products': {
                  product_id: {
                      'name': ...,
                      'ordered_qty': ...,  # qty on the original sale order
                      'invoiced_qty': ..., # qty actually invoiced (delivered)
                      'sale_qty': ...,     # invoiced_qty - return_qty (net sold)
                      'return_qty': ...,   # returned via stock or credit note
                      'ordered_amt': ...,
                      'sale_amt': ...,
                      'return_amt': ...,
                  }
              },
              'paid_amt': ...,
          }
        """
        wanted_products = self.product_ids.ids  # empty = all

        return_map = self._build_return_map(
            data['invoices'], data['returns'], wanted_products
        )

        result = defaultdict(lambda: defaultdict(lambda: {
            'partner_name': '',
            'invoice_name': '',
            'invoice_id': False,
            'invoice_date': False,
            'paid_amt': 0.0,
            'products': defaultdict(lambda: {
                'name': '',
                'ordered_qty': 0.0, 'invoiced_qty': 0.0,
                'return_qty': 0.0,  'sale_qty': 0.0,
                'ordered_amt': 0.0, 'sale_amt': 0.0, 'return_amt': 0.0,
            })
        }))

        for inv in data['invoices']:
            inv_date = inv.invoice_date
            pid = inv.partner_id.id
            result[inv_date][pid]['partner_name'] = inv.partner_id.name
            result[inv_date][pid]['invoice_name'] = inv.name
            result[inv_date][pid]['invoice_id']   = inv.id
            result[inv_date][pid]['invoice_date'] = inv_date
            result[inv_date][pid]['paid_amt']     = inv.amount_total - inv.amount_residual

            for line in inv.invoice_line_ids.filtered(
                lambda l: l.product_id and l.display_type not in ('line_section', 'line_note')
            ):
                if wanted_products and line.product_id.id not in wanted_products:
                    continue
                prod_id = line.product_id.id
                pdata   = result[inv_date][pid]['products'][prod_id]
                pdata['name'] = line.product_id.display_name

                # Invoiced qty = qty on this invoice line (what was billed)
                pdata['invoiced_qty'] += line.quantity
                pdata['ordered_amt']  += line.price_subtotal

                # Ordered qty = original SO qty (before any return/partial delivery)
                for sol in line.sale_line_ids:
                    pdata['ordered_qty'] += sol.product_uom_qty

                # Return qty from our combined map (stock + credit note)
                rkey = (inv.id, prod_id)
                ret_qty = return_map[rkey]['qty']
                ret_amt = return_map[rkey]['amt']

                pdata['return_qty'] = ret_qty
                pdata['return_amt'] = ret_amt

                unit_price = line.price_unit * (1 - line.discount / 100.0)
                pdata['sale_qty'] = max(pdata['invoiced_qty'] - ret_qty, 0)
                pdata['sale_amt'] = max(pdata['ordered_amt'] - ret_amt, 0)

        return result

    # ── Excel Builder ─────────────────────────────────────────────────────────
    def _build_excel(self, data):
        wb = openpyxl.Workbook()

        # Sheet 1: Invoice-wise detail (main report)
        ws_main = wb.active
        ws_main.title = "Return Report"
        self._write_main_sheet(ws_main, data)

        # Sheet 2: Product Summary
        ws_prod = wb.create_sheet("Product Summary")
        self._write_product_summary_sheet(ws_prod, data)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Styles (same palette as Collection Report) ────────────────────────────
    @staticmethod
    def _styles():
        thin = Side(style='thin', color='BDBDBD')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        thick_bottom = Border(
            left=thin, right=thin,
            top=thin, bottom=Side(style='medium', color='2E75B6')
        )
        return {
            'title':      Font(name='Arial', bold=True, size=14, color='FFFFFF'),
            'header':     Font(name='Arial', bold=True, size=9,  color='FFFFFF'),
            'sub_header': Font(name='Arial', bold=True, size=9,  color='FFFFFF'),
            'body':       Font(name='Arial', size=9),
            'body_bold':  Font(name='Arial', bold=True, size=9),
            'italic':     Font(name='Arial', italic=True, size=8, color='666666'),

            'title_fill':   PatternFill('solid', start_color='1F3864'),
            'header_fill':  PatternFill('solid', start_color='2E75B6'),
            'grp_fill':     PatternFill('solid', start_color='D6E4F0'),   # invoice group header
            'alt_fill':     PatternFill('solid', start_color='EBF3FB'),
            'white_fill':   PatternFill('solid', start_color='FFFFFF'),
            'total_fill':   PatternFill('solid', start_color='D6DCE4'),
            'return_fill':  PatternFill('solid', start_color='FFC7CE'),   # red tint
            'sale_fill':    PatternFill('solid', start_color='C6EFCE'),   # green tint
            'paid_fill':    PatternFill('solid', start_color='FFEB9C'),   # yellow tint

            'center': Alignment(horizontal='center', vertical='center', wrap_text=True),
            'left':   Alignment(horizontal='left',   vertical='center'),
            'right':  Alignment(horizontal='right',  vertical='center'),
            'border': border,
            'thick_bottom': thick_bottom,
        }

    # ── Sheet 1: Main Return Report ───────────────────────────────────────────
    def _write_main_sheet(self, ws, data):
        s = self._styles()
        agg = self._aggregate(data)

        # ── Collect all unique products (for dynamic columns) ────────────────
        all_products = {}   # product_id -> name
        for date_data in agg.values():
            for partner_data in date_data.values():
                for prod_id, pd in partner_data['products'].items():
                    all_products[prod_id] = pd['name']
        all_products = dict(sorted(all_products.items(), key=lambda x: x[1]))
        prod_list = list(all_products.items())   # [(id, name), ...]

        # ── Column layout ────────────────────────────────────────────────────
        # Fixed cols: Invoice Date | Shop Name | Invoice No. | Total Paid
        # Then per product: Ordered Qty | Sale Qty | Return Qty
        # Last col: Grand Total Return Qty

        FIXED = 4   # A..D
        cols_per_prod = 3  # Ordered / Sale / Return
        total_cols = FIXED + len(prod_list) * cols_per_prod + 1  # +1 summary

        # ── Title ────────────────────────────────────────────────────────────
        last_col_letter = get_column_letter(total_cols)
        ws.merge_cells(f'A1:{last_col_letter}1')
        ws['A1'] = f"RETURN REPORT  |  {self.period_label}"
        ws['A1'].font = s['title']
        ws['A1'].fill = s['title_fill']
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 28

        ws.merge_cells(f'A2:{last_col_letter}2')
        ws['A2'] = f"Generated: {date.today()}  |  Period: {self.date_from}  to  {self.date_to}"
        ws['A2'].font = s['italic']
        ws['A2'].alignment = s['center']
        ws.row_dimensions[2].height = 14

        # ── Row 3: Main column group headers ────────────────────────────────
        row3 = 3
        # Fixed group headers
        for col, label in enumerate(['Invoice Date', 'Shop Name', 'Invoice No.', 'Total Paid'], 1):
            ws.merge_cells(start_row=row3, start_column=col, end_row=row3+1, end_column=col)
            cell = ws.cell(row3, col, label)
            cell.font = s['header']
            cell.fill = s['header_fill']
            cell.alignment = s['center']
            cell.border = s['border']

        # Per-product group headers (merged across 3 sub-cols)
        for i, (prod_id, prod_name) in enumerate(prod_list):
            start_col = FIXED + 1 + i * cols_per_prod
            end_col   = start_col + cols_per_prod - 1
            ws.merge_cells(start_row=row3, start_column=start_col, end_row=row3, end_column=end_col)
            cell = ws.cell(row3, start_col, prod_name)
            cell.font = s['sub_header']
            cell.fill = PatternFill('solid', start_color='2E75B6') if i % 2 == 0 else PatternFill('solid', start_color='1F5F9B')
            cell.alignment = s['center']
            cell.border = s['border']

        # Grand return summary header
        summary_col = FIXED + len(prod_list) * cols_per_prod + 1
        ws.merge_cells(start_row=row3, start_column=summary_col, end_row=row3+1, end_column=summary_col)
        cell = ws.cell(row3, summary_col, 'Total Return Qty')
        cell.font = s['header']
        cell.fill = PatternFill('solid', start_color='C00000')
        cell.alignment = s['center']
        cell.border = s['border']

        # ── Row 4: Sub-headers for each product ─────────────────────────────
        row4 = 4
        # Fill merged fixed cells in row4
        for col in range(1, FIXED + 1):
            c = ws.cell(row4, col)
            c.fill = s['header_fill']
            c.border = s['border']

        for i, (prod_id, _) in enumerate(prod_list):
            start_col = FIXED + 1 + i * cols_per_prod
            sub_labels = ['SO Qty', 'Invoiced', 'Return']
            sub_fills  = [s['white_fill'], s['sale_fill'], s['return_fill']]
            for j, (lbl, sfill) in enumerate(zip(sub_labels, sub_fills)):
                cell = ws.cell(row4, start_col + j, lbl)
                cell.font = Font(name='Arial', bold=True, size=8, color='000000')
                cell.fill = sfill
                cell.alignment = s['center']
                cell.border = s['border']

        c = ws.cell(row4, summary_col)
        c.fill = PatternFill('solid', start_color='C00000')
        c.border = s['border']

        ws.row_dimensions[row3].height = 30
        ws.row_dimensions[row4].height = 16

        # ── Data Rows ────────────────────────────────────────────────────────
        data_row = 5
        grand_ordered = defaultdict(float)
        grand_sale    = defaultdict(float)
        grand_return  = defaultdict(float)
        grand_paid    = 0.0

        for inv_date in sorted(agg.keys()):
            for partner_id, pdata in sorted(agg[inv_date].items(), key=lambda x: x[1]['partner_name']):
                is_alt = (data_row % 2 == 0)
                row_fill = s['alt_fill'] if is_alt else s['white_fill']

                total_return_qty = sum(p['return_qty'] for p in pdata['products'].values())
                row_highlight = s['return_fill'] if total_return_qty > 0 else row_fill

                # Fixed cols
                fixed_vals = [
                    (inv_date,              s['center'], 'DD/MM/YYYY'),
                    (pdata['partner_name'], s['left'],   None),
                    (pdata['invoice_name'], s['left'],   None),
                    (pdata['paid_amt'],     s['right'],  '#,##0.00'),
                ]
                for col, (val, align, fmt) in enumerate(fixed_vals, 1):
                    cell = ws.cell(data_row, col, val)
                    cell.font = s['body']
                    cell.fill = row_fill
                    cell.alignment = align
                    cell.border = s['border']
                    if fmt:
                        cell.number_format = fmt

                grand_paid += pdata['paid_amt']

                # Product cols
                row_total_return = 0
                for i, (prod_id, _) in enumerate(prod_list):
                    start_col = FIXED + 1 + i * cols_per_prod
                    pd = pdata['products'].get(prod_id, {})
                    oq = pd.get('ordered_qty', 0)    # original SO qty
                    sq = pd.get('invoiced_qty', 0)   # invoiced/delivered qty
                    rq = pd.get('return_qty', 0)     # returned qty
                    row_total_return += rq
                    grand_ordered[prod_id] += oq
                    grand_sale[prod_id]    += sq
                    grand_return[prod_id]  += rq

                    for j, (val, fill_key) in enumerate([(oq, 'white_fill'), (sq, 'sale_fill'), (rq, 'return_fill')]):
                        cell = ws.cell(data_row, start_col + j, val if val else '')
                        cell.font = s['body']
                        cell.fill = s[fill_key] if val else row_fill
                        cell.alignment = s['center']
                        cell.border = s['border']
                        cell.number_format = '#,##0.##'

                # Summary col
                cell = ws.cell(data_row, summary_col, row_total_return if row_total_return else '')
                cell.font = s['body_bold']
                cell.fill = s['return_fill'] if row_total_return else row_fill
                cell.alignment = s['center']
                cell.border = s['border']

                data_row += 1

        # ── Grand Total Row ───────────────────────────────────────────────────
        ws.cell(data_row, 1, 'GRAND TOTAL').font = s['body_bold']
        ws.cell(data_row, 1).fill = s['total_fill']
        ws.cell(data_row, 1).alignment = s['left']
        ws.cell(data_row, 1).border = s['border']
        for col in range(2, FIXED + 1):
            c = ws.cell(data_row, col)
            c.fill = s['total_fill']
            c.border = s['border']
            c.font = s['body_bold']
        ws.cell(data_row, 4, grand_paid).number_format = '#,##0.00'
        ws.cell(data_row, 4).font = s['body_bold']
        ws.cell(data_row, 4).fill = s['total_fill']
        ws.cell(data_row, 4).alignment = s['right']
        ws.cell(data_row, 4).border = s['border']

        total_all_returns = 0
        for i, (prod_id, _) in enumerate(prod_list):
            start_col = FIXED + 1 + i * cols_per_prod
            vals_t = [grand_ordered[prod_id], grand_sale[prod_id], grand_return[prod_id]]
            total_all_returns += grand_return[prod_id]
            for j, val in enumerate(vals_t):
                cell = ws.cell(data_row, start_col + j, val if val else '')
                cell.font = s['body_bold']
                cell.fill = s['total_fill']
                cell.alignment = s['center']
                cell.border = s['border']
                cell.number_format = '#,##0.##'

        cell = ws.cell(data_row, summary_col, total_all_returns if total_all_returns else '')
        cell.font = s['body_bold']
        cell.fill = PatternFill('solid', start_color='C00000')
        cell.alignment = s['center']
        cell.border = s['border']
        ws.cell(data_row, summary_col).font = Font(name='Arial', bold=True, size=9, color='FFFFFF')

        # ── Column Widths ─────────────────────────────────────────────────────
        ws.column_dimensions['A'].width = 14  # date
        ws.column_dimensions['B'].width = 26  # shop
        ws.column_dimensions['C'].width = 18  # invoice no
        ws.column_dimensions['D'].width = 14  # paid
        for i in range(len(prod_list) * cols_per_prod + 1):
            ws.column_dimensions[get_column_letter(FIXED + 1 + i)].width = 10

        ws.freeze_panes = 'A5'
        ws.row_dimensions[data_row].height = 18

    # ── Sheet 2: Product Summary ──────────────────────────────────────────────
    def _write_product_summary_sheet(self, ws, data):
        s = self._styles()
        agg = self._aggregate(data)

        # Accumulate totals per product
        prod_totals = defaultdict(lambda: {
            'name': '', 'ordered_qty': 0.0, 'invoiced_qty': 0.0,
            'return_qty': 0.0, 'ordered_amt': 0.0, 'sale_amt': 0.0, 'return_amt': 0.0
        })
        for date_data in agg.values():
            for partner_data in date_data.values():
                for prod_id, pd in partner_data['products'].items():
                    prod_totals[prod_id]['name']          = pd['name']
                    prod_totals[prod_id]['ordered_qty']  += pd['ordered_qty']
                    prod_totals[prod_id]['invoiced_qty'] += pd['invoiced_qty']
                    prod_totals[prod_id]['return_qty']   += pd['return_qty']
                    prod_totals[prod_id]['ordered_amt']  += pd['ordered_amt']
                    prod_totals[prod_id]['sale_amt']     += pd['sale_amt']
                    prod_totals[prod_id]['return_amt']   += pd['return_amt']

        ws.merge_cells('A1:H1')
        ws['A1'] = f"PRODUCT RETURN SUMMARY  |  {self.period_label}"
        ws['A1'].font = s['title']
        ws['A1'].fill = s['title_fill']
        ws['A1'].alignment = s['center']
        ws.row_dimensions[1].height = 28

        headers = [
            'Product', 'SO Ordered Qty', 'Invoiced Qty', 'Return Qty',
            'Return %', 'Invoiced Amount', 'Net Sale Amount', 'Return Amount'
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(2, col, h)
            cell.font = s['header']
            cell.fill = s['header_fill']
            cell.alignment = s['center']
            cell.border = s['border']
        ws.row_dimensions[2].height = 20

        for i, (prod_id, pd) in enumerate(sorted(prod_totals.items(), key=lambda x: x[1]['name'])):
            row = i + 3
            ret_pct = (pd['return_qty'] / pd['invoiced_qty']) if pd['invoiced_qty'] else 0
            fill = s['alt_fill'] if i % 2 == 0 else s['white_fill']
            row_vals = [
                (pd['name'],          s['left'],   None),
                (pd['ordered_qty'],   s['center'], '#,##0.##'),
                (pd['invoiced_qty'],  s['center'], '#,##0.##'),
                (pd['return_qty'],    s['center'], '#,##0.##'),
                (ret_pct,             s['center'], '0.0%'),
                (pd['ordered_amt'],   s['right'],  '#,##0.00'),
                (pd['sale_amt'],      s['right'],  '#,##0.00'),
                (pd['return_amt'],    s['right'],  '#,##0.00'),
            ]
            for col, (val, align, fmt) in enumerate(row_vals, 1):
                cell = ws.cell(row, col, val)
                cell.font = s['body']
                cell.fill = s['return_fill'] if (col == 4 and pd['return_qty'] > 0) else fill
                cell.alignment = align
                cell.border = s['border']
                if fmt:
                    cell.number_format = fmt

        # Totals
        last_data = len(prod_totals) + 2
        total_row = last_data + 1
        ws.cell(total_row, 1, 'TOTAL').font = s['body_bold']
        ws.cell(total_row, 1).fill = s['total_fill']
        ws.cell(total_row, 1).alignment = s['left']
        ws.cell(total_row, 1).border = s['border']
        for col in range(2, 9):
            cell = ws.cell(total_row, col)
            cell.font = s['body_bold']
            cell.fill = s['total_fill']
            cell.border = s['border']
            col_letter = get_column_letter(col)
            if col != 5:
                cell.value = f'=SUM({col_letter}3:{col_letter}{last_data})'
                cell.number_format = '#,##0.##' if col in (2, 3, 4) else '#,##0.00'
                cell.alignment = s['center'] if col in (2, 3, 4) else s['right']
            else:
                cell.value = f'=IF(B{total_row}=0,0,D{total_row}/B{total_row})'
                cell.number_format = '0.0%'
                cell.alignment = s['center']

        widths = [30, 18, 15, 15, 12, 16, 16, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:H{last_data}'
